from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from holywar.data.models import CardDefinition


def _to_int(value: object) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def load_cards_from_xlsx(xlsx_path: str | Path) -> list[CardDefinition]:
    wb = load_workbook(xlsx_path, data_only=True)
    cards: list[CardDefinition] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
            name = str(row[0]).strip() if row[0] is not None else ""
            if not name:
                continue
            card_type = str(row[1]).strip() if row[1] is not None else ""
            crosses = str(row[2]).strip() if row[2] is not None else ""
            faith = _to_int(row[3])
            strength = _to_int(row[4])
            effect_text = str(row[5]).strip() if row[5] is not None else ""
            expansion = str(row[6]).strip() if row[6] is not None else sheet_name
            is_token = card_type.lower() == "token"
            cards.append(
                CardDefinition(
                    name=name,
                    card_type=card_type,
                    crosses=crosses,
                    faith=faith,
                    strength=strength,
                    effect_text=effect_text,
                    expansion=expansion,
                    is_token=is_token,
                )
            )
    return cards


def write_cards_json(cards: list[CardDefinition], out_path: str | Path) -> Path:
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    payload = [c.to_dict() for c in cards]
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return out


def load_cards_json(path: str | Path) -> list[CardDefinition]:
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    return [CardDefinition.from_dict(item) for item in data]